import glob
import json
import logging
import os
import shutil
import tempfile
import time
import traceback
from dataclasses import dataclass
from datetime import datetime
from functools import wraps
from io import BytesIO
from pathlib import Path

import openpyxl
import pythoncom
import requests

import xlwings as xw

from flask import Response
from werkzeug.datastructures import FileStorage

from app_log import debug, error
from database import Database
from msgbox_listener import MsgBoxListener
from utils import is_file_write_locked, today_as_string

CUSTOM_DATA_BOOK = "custom_data.xlsx"
TOOL12_BOOK = "Tool 1.2.xlsx"
AUDIT_BOOK = "audit.xlsx"
CLIP_BOOK = "CLIP.xlsx"
CLIP_SHEET = "CLIP"


PROJECT_NAME_COL = "ProjectName"
PROJECT_PATH_COL = "ProjectPath"
CLIENT_COL = "Client"

PROPOSAL_DIR = "Proposal"
SCOPE_DIR = "Scope"

NO_DOWNLOAD = "<:NO_DOWNLOAD:>"
USE_UPLOAD_NAME = ""
DEFAULT_RESPONSE = "OK"

FILE_IN_USE_ERR = "File access error. Please try again."

@dataclass
class MockRequest:
    files: dict
    form: dict


@dataclass
class VBAError(Exception):
    message: str


@dataclass
class APPError(Exception):
    message: str


class XWApp():
    def __init__(self):
        self.excel = xw.App()

    def __enter__(self):
        return self.excel

    def __exit__(self, *args):
        debug(f"Exiting Excel.")

        for book in self.excel.books:
            try:
                book.close()
            except: pass

        self.excel.kill()

        debug(f"Excel process has exited.")


class RequestAsProject:
    def __init__(self, request):
        self._request = request

        project_rows = request.json

        if len(project_rows):
            project = project_rows[0]
            self.name = project[PROJECT_NAME_COL]
            self.path = project[PROJECT_PATH_COL]
            self.proposal_folder = os.path.join(self.path, PROPOSAL_DIR)

    def find_latest_audit(self):
        return self.get_latest_file_from(self.proposal_folder, AUDIT_BOOK)

    def find_latest_tool(self):
        return self.get_latest_file_from(self.proposal_folder, TOOL12_BOOK)

    def get_latest_file_from(self, find_in_folder, file_to_find):
        max_time = 0
        latest_audit_path = None

        for root, dirs, files in os.walk(find_in_folder):
            for name in files:
                if name == file_to_find:
                    audit_path = os.path.join(root, name)
                    file_time = os.path.getmtime(audit_path)

                    if file_time > max_time:
                        max_time = file_time
                        latest_audit_path = audit_path

        return latest_audit_path


def reraise_if_enabled(e, kwargs):
    if kwargs.get("raise_errors", False):
        raise e


def catch_errors(request_logic):
    @wraps(request_logic)
    def wrapped(*args, **kwargs):
        err = None
        result = None

        try:
            result = request_logic(*args, **kwargs)
        except (APPError, VBAError) as e:
            reraise_if_enabled(e, kwargs)

            err = get_error_json(e)
            err["display"] = True
        except Exception as e:
            reraise_if_enabled(e, kwargs)

            err = get_error_json(e)
            print(traceback.format_exc())
            error(traceback.format_exc())

        if result:
            return result, 200
        else:
            return err, 500

    return wrapped


def get_error_json(err):
    return {
        "error": True,
        "display": False,
        "message": str(err)
    }


def macro_with_new_excel_app(out_ext="xlsx", **extras):
    def wrapper(request_logic):
        @wraps(request_logic)
        @catch_errors
        def wrapped(self, request, **kwargs):
            pythoncom.CoInitialize()

            result = None
            tmp_dir = tempfile.mkdtemp()
            timestamp = round(time.time() * 1000)
            in_file_name = f"{timestamp}_in.xlsx"
            in_file_path = os.path.join(tmp_dir, in_file_name)
            out_file_path = os.path.join(tmp_dir, f"{timestamp}_out.{out_ext}")

            debug(f"Executing Excel macro.")
            debug(f"Temp dir: {tmp_dir}")

            if not self._debug:
                msgbox_listener = MsgBoxListener(2)
                msgbox_listener.start()

            try:
                has_content = "content" in request.files

                if has_content:
                    request.files["content"].save(in_file_path)

                download_name = None

                with XWApp() as xlapp:
                    debug(f"Started Excel with PID: {xlapp.pid}")

                    xlapp.display_alerts = False
                    in_book = xlapp.books.open(in_file_path, update_links=True) if has_content else None

                    debug(f"Opened input book: {in_file_path}")

                    macro_book_path = extras.get("book", None)
                    macro_book = None

                    debug(f"Opened macro book: {macro_book_path}")

                    if macro_book_path:
                        macro_book_path = self.get_macro_path(macro_book_path)
                        macro_book = xlapp.books.open(macro_book_path, read_only=True)

                    debug(f"Invoking the macro.")

                    download_name = request_logic(self, macro_book, in_book, out_file_path,
                                                  request=request, xlapp=xlapp, tmp_dir=tmp_dir, **kwargs)

                    debug(f"The macro has finished.")
                    try:
                        in_book and in_book.close()
                        macro_book and macro_book.close()
                    except: pass

                if isinstance(download_name, str):
                    if download_name == NO_DOWNLOAD:
                        result = DEFAULT_RESPONSE
                    else:
                        with open(out_file_path, mode="rb") as file:
                            file_name = download_name or request.form.get("file_name", "")
                            result = Response(file.read(), headers={"x-file-name": file_name})
                else:
                    result = DEFAULT_RESPONSE

            finally:
                debug("Finally block")

                time.sleep(1)
                try:
                    shutil.rmtree(tmp_dir)
                except Exception as e:
                    pass

                if not self._debug:
                    msgbox_listener.stop()

                    if msgbox_listener.file_in_use_dialog:
                        raise APPError(FILE_IN_USE_ERR)

            return result

        return wrapped

    return wrapper


class ExcelProcessor:

    def __init__(self, flask_app):
        self._flask_app = flask_app
        self._db = Database(flask_app.config["DATABASE_PATH"])
        self._debug = not not os.getenv("FLASK_DEBUG")
        self.XL_DATABASE = flask_app.config["DATABASE_PATH"]
        self.XL_TEMPLATE_PATH = flask_app.config["XL_TEMPLATE_PATH"]

    def get_app_path(self, path):
        return f"{self._flask_app.root_path}/{path}"

    def get_macro_path(self, path):
        return self.get_app_path(f"excel/macros/{path}")

    def get_template_path(self, path):
        return f"{self.XL_TEMPLATE_PATH}/{path}"

    def get_local_db_path(self, start_folder):
        custom_db_path = os.path.join(start_folder, CUSTOM_DATA_BOOK)

        if os.path.exists(custom_db_path):
            return custom_db_path
        else:
            try:
                if os.path.exists(start_folder) \
                        and (os.path.samefile(start_folder, self._flask_app.config["PROJECTS_BASE_PATH"])
                             or os.path.samefile(start_folder, os.path.splitdrive(start_folder)[0])):  # root drive
                    return self.XL_DATABASE
                else:
                    return self.get_local_db_path(os.path.dirname(start_folder))
            except Exception as e:
                return self.XL_DATABASE

    def execute_macro(self, book, macro_name, *args):
        macro = book.macro(macro_name)
        err = macro(self._debug, *args)

        if err:
            err = json.loads(err)
            raise VBAError("VBA Error: " + err["message"])

    def create_mock_flask_request(self, file_path, file_name):
        with open(file_path, "rb") as fp:
            file = FileStorage(BytesIO(fp.read()))

            return MockRequest(
                {"content": file},
                {"file_name": file_name}
            )

    @catch_errors
    def create_new_project_from_raw_report(self, request):
        audit_bytes = BytesIO(request.files["content"].read())
        wb = openpyxl.load_workbook(filename=audit_bytes)
        ws = wb.worksheets[0]

        customer_name = ws.cell(1, 1).value.strip()
        customer_address = ws.cell(2, 1).value.strip()
        customer_address = (" ".join(customer_address.split(",")[0].split(" ")[:-1])).strip()

        projects_folder = request.form.get("projects_folder", None)

        if not projects_folder:
            db_config = self._db.get_config()
            projects_folder = db_config[Database.DEFAULT_PROJECTS_PATH]

        project_folder = os.path.join(projects_folder, customer_name, customer_address)
        proposal_folder = os.path.join(project_folder, PROPOSAL_DIR)
        project_version = 1 if not os.path.exists(project_folder) else self.count_tool_books(proposal_folder)
        proposal_folder = os.path.join(proposal_folder, f"v{project_version}")
        audit_book = os.path.join(proposal_folder, request.form["file_name"])
        db_path = self.get_local_db_path(proposal_folder)

        Path(proposal_folder).mkdir(parents=True, exist_ok=True)

        with open(audit_book, "wb") as out_file:
            out_file.write(audit_bytes.getbuffer())

        request.files["content"].stream.seek(0)
        result, status = self.process_audit_report(request, db_path=db_path, raise_errors=True)

        if status != 500:
            processed_audit_book = os.path.join(proposal_folder, AUDIT_BOOK)
            with open(processed_audit_book, "wb") as out_file:
                out_file.write(result.data)

        projects = self.append_project_to_db(customer_address, customer_name, project_folder)

        return Response(json.dumps(projects), mimetype="application/json")

    def count_tool_books(self, proposal_folder):
        tool_count = 0

        for root, dirs, files in os.walk(proposal_folder):
            for name in files:
                if name.lower() == TOOL12_BOOK.lower():
                    tool_count += 1

        return tool_count + 1

    def append_project_to_db(self, project_name, customer, project_folder):
        projects, _ = self._db.get_projects()
        existing_project = self._db.find_project(projects, project_name)

        if not existing_project:
            projects.append({
                PROJECT_NAME_COL: project_name,
                CLIENT_COL: customer,
                PROJECT_PATH_COL: project_folder
            })

            self._db.write_projects(projects)

            wb_db = openpyxl.load_workbook(self.XL_DATABASE)
            ws_projects = wb_db[Database.PROJECTS_SHEET]

            for i in range(ws_projects.max_row, 0, -1):
                project_cell = ws_projects.cell(i, 1)

                if project_cell.value:
                    project_cell.hyperlink = f"file:///{project_folder}"
                    project_cell.style = "Hyperlink"
                    wb_db.save(self.XL_DATABASE)
                    break

        return projects

    @catch_errors
    def reprocess_audit_report(self, request, reset=False):
        project = RequestAsProject(request)
        latest_audit_book = project.find_latest_audit()

        if latest_audit_book:
            if is_file_write_locked(latest_audit_book):
                raise APPError(FILE_IN_USE_ERR)

            db_path = self.get_local_db_path(os.path.dirname(latest_audit_book))

            if reset:
                # remove all worksheets except the first
                wb = openpyxl.load_workbook(latest_audit_book)

                if len(wb.worksheets) > 1:
                    for n in wb.get_sheet_names()[1:]:
                        wb.remove_sheet(wb.get_sheet_by_name(n))

                    wb.save(latest_audit_book)

            mock_request = self.create_mock_flask_request(latest_audit_book, AUDIT_BOOK)
            result, status = self.process_audit_report(mock_request, db_path=db_path, raise_errors=True)

            if status != 500:
                with open(latest_audit_book, "wb") as out_file:
                    out_file.write(result.data)
        else:
            raise APPError("Missing audit.xlsx")

        return DEFAULT_RESPONSE

    @macro_with_new_excel_app(book="audit/macro_audit.xlsm")
    def process_audit_report(self, macro_book, in_book, out_file_path, **kwargs):
        db_path = kwargs.get("db_path", False) or self.XL_DATABASE

        in_book.activate()

        self.execute_macro(macro_book, "ProcessAuditWithDB", db_path, out_file_path)

        return USE_UPLOAD_NAME  # use the request file name as the download name

    @macro_with_new_excel_app(book="proposal/macro_proposal.xlsm")
    def create_proposal_pdf(self, macro_book, in_book, out_file_path, **kwargs):
        project = RequestAsProject(kwargs["request"])
        latest_audit_book = project.find_latest_audit()

        if latest_audit_book:
            debug(f"Proposal project path: {project.path}")

            db_path = self.get_local_db_path(os.path.dirname(latest_audit_book))
            tool_12_form_path = self.get_template_path(f"proposal/{TOOL12_BOOK}")
            audit_directory = os.path.dirname(latest_audit_book)
            tool_12_instance_path = os.path.join(audit_directory, TOOL12_BOOK)
            logo_path = ""

            if not os.path.exists(tool_12_instance_path):
                shutil.copy(tool_12_form_path, tool_12_instance_path)
                debug(f"Copied Tool 1.2 template: {tool_12_instance_path}")

                # sleep to reduce the probability of file in use error
                time.sleep(0.5)

                logo_path = os.path.join(project.path, "logo.*")
                logo_path = next(iter(glob.glob(logo_path)), "")

                if not logo_path:
                    logo_path = os.path.join(audit_directory, "logo.*")
                    logo_path = next(iter(glob.glob(logo_path)), "")

            # DEBUG: copy audit to the tmp dir to check for the Dropbox-related error
            try:
                tmp_audit_path = os.path.join(kwargs["tmp_dir"], AUDIT_BOOK)
                shutil.copy(latest_audit_book, tmp_audit_path)
                latest_audit_book = tmp_audit_path
            except Exception as e:
                raise APPError(str(e))

            debug(f"Copied audit.xslx: {tmp_audit_path}, the file size is: {os.stat(tmp_audit_path).st_size}")

            if is_file_write_locked(tool_12_instance_path):
                raise APPError(FILE_IN_USE_ERR)

            self.execute_macro(macro_book, "CreateProposalPDFWithTemplates", db_path, latest_audit_book,
                               tool_12_instance_path, audit_directory, kwargs["tmp_dir"], logo_path, False)
        else:
            raise APPError("Missing audit.xlsx")

    @macro_with_new_excel_app(book="proposal/macro_proposal.xlsm")
    def create_proposal_custom(self, macro_book, in_book, out_file_path, **kwargs):
        audit_book_path = in_book.fullname
        tool_12_form_path = self.get_template_path(f"proposal/{TOOL12_BOOK}")
        audit_directory = os.path.dirname(out_file_path)
        tool_12_instance_path = out_file_path
        logo_path = ""

        shutil.copy(tool_12_form_path, tool_12_instance_path)
        # necessary for references to the audit.xlsx in the same dir as Tool 1.2
        shutil.copy(audit_book_path, os.path.join(audit_directory, AUDIT_BOOK))

        if is_file_write_locked(tool_12_instance_path):
            raise APPError(FILE_IN_USE_ERR)

        self.execute_macro(macro_book, "CreateProposalPDFWithTemplates", self.XL_DATABASE, audit_book_path,
                           tool_12_instance_path, audit_directory, kwargs["tmp_dir"], logo_path, True)

        return TOOL12_BOOK

    @macro_with_new_excel_app(book="photos_check_list/macro_pcl.xlsm")
    def create_photos_check_list(self, macro_book, in_book, out_file_path, **kwargs):
        self.execute_macro(macro_book, "CreatePhotosCheckListWithDB", self.XL_DATABASE, in_book.name)

        macro_book.save(out_file_path)

        return f"Photos Check List - {today_as_string()}.xlsx"

    @macro_with_new_excel_app(book="scope/macro_scope.xlsm", out_ext="pdf")
    def create_scope_pdf(self, macro_book, in_book, out_file_path, **kwargs):
        proposal_form_path = self.get_template_path("scope/Scope From Proposal.xlsx")
        scope_form_path = self.get_template_path("scope/Scope v2 - XLS.xlsx")
        word_form_path = self.get_template_path("scope/Scope v2.docx")
        scope_xls_path = kwargs.get("scope_xls_path", "")
        db_path = kwargs.get("db_path", False) or self.XL_DATABASE

        # the macro saves intermediate XLSX book by the path in scope_xls_path
        self.execute_macro(macro_book, "CreateScopePDFWithTemplates", db_path, in_book.name, proposal_form_path,
                      scope_form_path, word_form_path, scope_xls_path, out_file_path)

        return f"Scope - {today_as_string()}.pdf"

    @catch_errors
    def create_project_scope_pdf(self, request):
        project = RequestAsProject(request)

        latest_tool_book = project.find_latest_tool()

        if latest_tool_book:
            db_path = self.get_local_db_path(os.path.dirname(latest_tool_book))
            scope_folder = os.path.join(project.path, SCOPE_DIR)
            Path(scope_folder).mkdir(parents=True, exist_ok=True)

            file_name_base = f"{project.name} - Scope - {today_as_string()}"
            scope_xls_path = os.path.join(scope_folder, file_name_base + ".xlsx")
            scope_pdf_path = os.path.join(scope_folder, file_name_base + ".pdf")
            mock_request = self.create_mock_flask_request(latest_tool_book, TOOL12_BOOK)
            result, status = self.create_scope_pdf(mock_request, scope_xls_path=scope_xls_path, db_path=db_path,
                                                   raise_errors=True)

            if status != 500:
                with open(scope_pdf_path, "wb") as out_file:
                    out_file.write(result.data)
        else:
            raise APPError("Missing Tool 1.2.xlsx")

        return NO_DOWNLOAD

    @macro_with_new_excel_app(book="clip/macro_clip.xlsm")
    def create_clip(self, macro_book, in_book, out_file_path, **kwargs):
        project = RequestAsProject(kwargs["request"])
        latest_audit_book = project.find_latest_audit()

        if latest_audit_book:
            db_path = self.get_local_db_path(os.path.dirname(latest_audit_book))

            #clip_path = self.download_clip_template(kwargs["tmp_dir"])
            clip_template_path = self.get_template_path(f"clip/{CLIP_BOOK}")
            shutil.copy(clip_template_path, kwargs["tmp_dir"])
            clip_path = os.path.join(kwargs["tmp_dir"], CLIP_BOOK)

            clip_output_folder = os.path.join(project.path, "CLIP")
            Path(clip_output_folder).mkdir(parents=True, exist_ok=True)

            self.execute_macro(macro_book, "CreateCLIPWithTemplates", db_path, latest_audit_book, clip_path,
                               clip_output_folder, project.name)
        else:
            raise APPError("Missing audit.xlsx")

        return NO_DOWNLOAD

    def download_clip_template(self, dest_folder):
        clip_url = "https://www.ladwp.com/cs/idcplg?IdcService=GET_FILE&dDocName=OPLADWPCCB411121&RevisionSelectionMethod=LatestReleased"
        clip_path = os.path.join(dest_folder, CLIP_BOOK)

        try:
            response = requests.get(clip_url)

            if response.status_code == 200:
                with open(clip_path, "wb") as clip_file:
                    clip_file.write(response.content)
            else:
                raise APPError("")

        except Exception as e:
            clip_template_path = self.get_template_path(f"clip/{CLIP_BOOK}")
            shutil.copy(clip_template_path, dest_folder)
            traceback.print_exc()
            #raise APPError("Error downloading CLIP template.")

        return clip_path

    @macro_with_new_excel_app(book="materials/macro_materials.xlsm")
    def create_materials(self, macro_book, in_book, out_file_path, **kwargs):
        project = RequestAsProject(kwargs["request"])
        clip_folder = os.path.join(project.path, "CLIP")
        clip_excel_books_mask = os.path.join(clip_folder, "*.xlsx")
        excel_books = glob.glob(clip_excel_books_mask)

        if len(excel_books):
            combined = kwargs.get("combined", False)
            db_path = self.get_local_db_path(clip_folder)
            materials_template = self.get_template_path("materials/Materials.xlsx")
            materials_folder = os.path.join(project.path, "Materials")
            Path(materials_folder).mkdir(parents=True, exist_ok=True)

            self.execute_macro(macro_book, "CreateMaterialsWithTemplates", db_path, materials_template, project.path,
                          project.name, combined)
        else:
            raise APPError("Missing CLIP books.")

